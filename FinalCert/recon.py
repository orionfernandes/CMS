import pandas as pd
import os
import glob
import shutil
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.exceptions import InvalidFileException

# --- Configuration Variables ---
CHASE_FOLDER = r'CHASE FOLDER PATH'
DB_FOLDER = r'DB FOLDER PATH'
SUMMARY_REPORT_PATH = r'SUMMARY REPORT GENERATION PATH'
IMPORT_FOLDER_PATH = r'IMPORT REPORT GENERATION PATH' 
ACCUMULATIVE_BASE_DIR = r'ACCUMULATIVE CLEARS GENERATION PATH'
BACKUP_DIR = os.path.join(ACCUMULATIVE_BASE_DIR, 'Backups')

os.makedirs(ACCUMULATIVE_BASE_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)
os.makedirs(IMPORT_FOLDER_PATH, exist_ok=True)

def format_all_dates(df):
    """Ensures dates are formatted as mm/dd/yyyy strings for Excel consistency."""
    for col in df.columns:
        if any(kw in str(col).lower() for kw in ['date', 'settlement']):
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%m/%d/%Y')
    return df

def get_target_files(folder_path, prefix, today_date):
    search_pattern = os.path.join(folder_path, f"{prefix}*.xlsx")
    all_files = glob.glob(search_pattern)
    valid_files = []
    for file in all_files:
        basename = os.path.basename(file)
        date_str = basename.replace(prefix, "").replace(".xlsx", "")
        try:
            f_date = datetime.strptime(date_str, '%b.%d').replace(year=today_date.year)
            if f_date > today_date + timedelta(days=30): f_date = f_date.replace(year=today_date.year - 1)
            valid_files.append((f_date, file))
        except ValueError: continue
    
    valid_files.sort(key=lambda x: x[0])
    if not valid_files: return None, None
    expected_today = f"{prefix}{today_date.strftime('%b.%d')}.xlsx"
    today_file, prev_file = None, None
    for i, (f_date, f_path) in enumerate(valid_files):
        if os.path.basename(f_path) == expected_today:
            today_file = f_path
            if i > 0: prev_file = valid_files[i-1][1]
            break
    return today_file, prev_file

def get_adds_and_clears(today_file, previous_file, sheet_name=0):
    try:
        df_today = pd.read_excel(today_file, sheet_name=sheet_name)
        df_prev = pd.read_excel(previous_file, sheet_name=sheet_name)
        df_today.columns = df_today.columns.str.strip()
        df_prev.columns = df_prev.columns.str.strip()
        df_today['ID_KEY'] = df_today['Unique ID'].astype(str).str.strip()
        df_prev['ID_KEY'] = df_prev['Unique ID'].astype(str).str.strip()
        
        today_ids = set(df_today['ID_KEY'].dropna())
        prev_ids = set(df_prev['ID_KEY'].dropna())
        
        adds_df = df_today[df_today['ID_KEY'].isin(today_ids - prev_ids)].copy()
        clears_df = df_prev[df_prev['ID_KEY'].isin(prev_ids - today_ids)].copy()
        
        # Calculate DQ category counts from today's file
        dq_91_180, dq_0_90, dq_3_yr = 0, 0, 0
        if 'Pool Age Status' in df_today.columns:
            pool_status = df_today['Pool Age Status'].astype(str).str.strip()
            dq_91_180 = len(pool_status[pool_status == 'DQ 91-180'])
            dq_0_90 = len(pool_status[pool_status == 'DQ 0-90'])
            dq_3_yr = len(pool_status[pool_status == '3yr DQ'])

        adds_df.drop(columns=['ID_KEY'], inplace=True, errors='ignore')
        clears_df.drop(columns=['ID_KEY'], inplace=True, errors='ignore')
        
        return len(adds_df), len(clears_df), adds_df, clears_df, dq_91_180, dq_0_90, dq_3_yr
    except Exception as e:
        print(f"Error comparing files: {e}")
        return 0, 0, pd.DataFrame(), pd.DataFrame(), 0, 0, 0

def update_accumulative_clears(clears_df, sheet_name, cleared_date_obj):
    if clears_df.empty: return
    cleared_date_str = cleared_date_obj.strftime('%m/%d/%Y')
    month_name = cleared_date_obj.strftime('%B_%Y') 
    file_path = os.path.join(ACCUMULATIVE_BASE_DIR, f"FC_Clear_Accumulative_{month_name}.xlsx")

    if os.path.exists(file_path):
        shutil.copy2(file_path, os.path.join(BACKUP_DIR, f"Backup_{os.path.basename(file_path)}"))

    clears_df = clears_df.copy()
    clears_df.insert(0, 'Cleared Date', cleared_date_str)
    clears_df = format_all_dates(clears_df)

    if not os.path.exists(file_path):
        wb = Workbook(); wb.active.title = "Chase"; wb.create_sheet("DB"); wb.save(file_path)
    
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames: wb.create_sheet(sheet_name)
    ws = wb[sheet_name]

    existing_records = set()
    if ws.max_row > 1:
        headers = [str(cell.value) for cell in ws[1]]
        try:
            date_idx = headers.index('Cleared Date')
            id_idx = headers.index('Unique ID')
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[date_idx] and row[id_idx]:
                    existing_records.add((str(row[date_idx]), str(row[id_idx])))
        except ValueError: pass

    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        for col_num, header in enumerate(clears_df.columns, 1):
            ws.cell(1, col_num, header).font = Font(bold=True)

    added_count = 0
    for row_dict in clears_df.to_dict('records'):
        record_key = (str(row_dict['Cleared Date']), str(row_dict['Unique ID']))
        if record_key not in existing_records:
            ws.append(list(row_dict.values()))
            added_count += 1
    
    wb.save(file_path)
    print(f"Logged {added_count} new clears to {sheet_name} ({month_name} file).")

def update_todays_clears_tab(c_clears_df, d_clears_df, cleared_date_obj):
    """Creates a snapshot tab of all combined clears for the current run."""
    month_name = cleared_date_obj.strftime('%B_%Y')
    file_path = os.path.join(ACCUMULATIVE_BASE_DIR, f"FC_Clear_Accumulative_{month_name}.xlsx")

    # If the file hasn't been created yet, exit (meaning 0 clears today)
    if not os.path.exists(file_path):
        return

    c_df = c_clears_df.copy()
    d_df = d_clears_df.copy()

    # Prep Chase Data
    if not c_df.empty:
        c_df.insert(0, 'Custodian', 'Chase')
        c_df.insert(0, 'Cleared Date', cleared_date_obj.strftime('%m/%d/%Y'))
        c_df = format_all_dates(c_df)
        
    # Prep DB Data
    if not d_df.empty:
        d_df.insert(0, 'Custodian', 'DB')
        d_df.insert(0, 'Cleared Date', cleared_date_obj.strftime('%m/%d/%Y'))
        d_df = format_all_dates(d_df)

    # Combine them
    combined_clears = pd.concat([c_df, d_df], ignore_index=True)

    wb = load_workbook(file_path)
    sheet_name = "Today's Clears"

    # Delete the old sheet if it exists so we have a fresh daily snapshot
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)

    if not combined_clears.empty:
        # Write Headers
        for col_num, header in enumerate(combined_clears.columns, 1):
            ws.cell(1, col_num, header).font = Font(bold=True)

        # Write Data
        for row in combined_clears.values.tolist():
            ws.append(row)
            
    wb.save(file_path)
    print(f"Updated '{sheet_name}' tab with {len(combined_clears)} total combined clears.")

def format_adds_for_import(adds_df, client_type, today_date_str):
    headers = [
        'Unique ID', 'Loan Number', 'Collateral Key', 'Settlement Date', 'Pool ID', 'Loan Status', 
        'Repool Investor', 'Prior Loan Number', 'Balance Principal', 'Responsible Party', 
        'Document Exception Status', 'Doctype Code', 'Doctype Descr', 'Doc Notation', 
        'Doc Status', 'Question Code', 'Question Descr', 'Question Notation', 
        'Exception Date', 'Clients Deal Name', 'Inv_#', 'Inv_Name'
    ]
    output_df = pd.DataFrame(columns=headers)
    if adds_df.empty: return output_df
    adds_df = format_all_dates(adds_df)

    if client_type == 'Chase':
        output_df['Unique ID'] = adds_df.get('Unique ID')
        output_df['Loan Number'] = adds_df.get('Collateral Key')
        output_df['Collateral Key'] = adds_df.get('Collateral Key')
        output_df['Settlement Date'] = adds_df.get('Settlement Date')
        output_df['Pool ID'] = adds_df.get('Pool Number')
        output_df['Loan Status'] = adds_df.get('Loan Status Current')
        output_df['Repool Investor'] = adds_df.get('Investor Name')
        output_df['Prior Loan Number'] = adds_df.get('Alternate Id')
        output_df['Balance Principal'] = adds_df.get('UPB') 
        output_df['Doc Status'] = adds_df.get('Doc Status')
        output_df['Inv_#'] = adds_df.get('Inv Num')
        output_df['Inv_Name'] = adds_df.get('Investor Name')
        output_df['Clients Deal Name'] = 'Chase'
    elif client_type == 'DB':
        output_df['Unique ID'] = adds_df.get('Unique ID')
        output_df['Loan Number'] = adds_df.get('Investor Collateral Key')
        output_df['Collateral Key'] = adds_df.get('Collateral Key')
        output_df['Settlement Date'] = adds_df.get('Settlement Date')
        output_df['Pool ID'] = adds_df.get('DB Pool Key')
        output_df['Loan Status'] = adds_df.get('Loan Status Current')
        output_df['Repool Investor'] = adds_df.get('Repool Investor')
        output_df['Prior Loan Number'] = adds_df.get('Prior Loan Number')
        output_df['Balance Principal'] = adds_df.get('Balance Principal')
        output_df['Doc Status'] = adds_df.get('Document Condition Code')
        output_df['Inv_#'] = adds_df.get('Investor Number')
        output_df['Inv_Name'] = adds_df.get('Investor Name')
        output_df['Clients Deal Name'] = 'DB'

    for field in ['Doctype Code', 'Doctype Descr', 'Doc Notation', 'Question Code', 'Question Descr', 'Question Notation']:
        output_df[field] = adds_df.get(field)
    
    # ⚠️ MOVED TO BOTTOM: Assign constants AFTER the dataframe has been expanded 
    # to fit the proper amount of rows from the .get() methods above.
    output_df['Responsible Party'] = 'Final Certification'
    output_df['Document Exception Status'] = 'Analyst Pending Review'
    output_df['Exception Date'] = today_date_str
    
    return output_df

def setup_summary_sheet_headers(ws, sheet_name):
    """Configures the precise multi-row layout of the Summary sheet tabs based on the image."""
    # Row 1: Master Header
    ws.cell(1, 1, f'{sheet_name}')
    ws.merge_cells('A1:G1')

    # Row 2: Sub-Headers
    ws.cell(2, 1, 'Date')
    ws.cell(2, 2, 'Adds')
    ws.cell(2, 3, 'Clears')
    ws.cell(2, 4, 'Exception Count')
    ws.cell(2, 5, 'DQ')

    # Row 3: DQ Sub-Buckets
    ws.cell(3, 5, '91-180')
    ws.cell(3, 6, '0-90')
    ws.cell(3, 7, '3 Year')

    # Vertical Merges for columns that don't split
    ws.merge_cells('A2:A3')
    ws.merge_cells('B2:B3')
    ws.merge_cells('C2:C3')
    ws.merge_cells('D2:D3')
    
    # Horizontal Merge for DQ
    ws.merge_cells('E2:G2')

    # Alignment and Font for headers
    for row in [1, 2, 3]:
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)

def update_individual_summary_tab(wb, sheet_name, today_str, adds, clears, dq_91, dq_0, dq_3):
    """Updates an individual sheet inside the Summary Workbook."""
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        setup_summary_sheet_headers(ws, sheet_name)
    else:
        ws = wb[sheet_name]
        if ws.max_row == 1 and ws.cell(1,1).value is None:
            setup_summary_sheet_headers(ws, sheet_name)

    # Deduplicate: Remove today's row if it already exists (Data starts at Row 4 now)
    for row_idx in range(4, ws.max_row + 1):
        if str(ws.cell(row_idx, 1).value) == today_str:
            ws.delete_rows(row_idx)
            break
            
    # Calculate running exception count
    last_row = ws.max_row
    # We check if last_row >= 4 because row 1, 2, and 3 are headers
    prev_exc = int(ws.cell(last_row, 4).value or 0) if last_row >= 4 else 0
    curr_exc = prev_exc + adds - clears

    # Append new data
    ws.append([today_str, adds, clears, curr_exc, dq_91, dq_0, dq_3])

def main():
    today = datetime.now()
    today_str = today.strftime('%m/%d/%Y')
    cleared_date = today - timedelta(days=3 if today.weekday() == 0 else 1)

    print(f"--- Process Start: {today_str} ---")

    # 1. Compare Files and Count DQs
    c_today, c_prev = get_target_files(CHASE_FOLDER, 'FCERT_Pipeline_Loan_Detail_', today)
    if c_today and c_prev:
        c_adds, c_clears, c_adds_df, c_clears_df, c_91, c_0, c_3 = get_adds_and_clears(c_today, c_prev, 'Loan Detail')
    else:
        c_adds, c_clears, c_adds_df, c_clears_df, c_91, c_0, c_3 = (0,0,pd.DataFrame(),pd.DataFrame(),0,0,0)

    d_today, d_prev = get_target_files(DB_FOLDER, 'DB_FCERT_Pipeline_', today)
    if d_today and d_prev:
        d_adds, d_clears, d_adds_df, d_clears_df, d_91, d_0, d_3 = get_adds_and_clears(d_today, d_prev)
    else:
        d_adds, d_clears, d_adds_df, d_clears_df, d_91, d_0, d_3 = (0,0,pd.DataFrame(),pd.DataFrame(),0,0,0)

    # 2. Update New Tabbed Summary Report
    print("Updating Separate Summary Tabs...")
    if not os.path.exists(SUMMARY_REPORT_PATH):
        wb = Workbook()
        if 'Sheet' in wb.sheetnames: wb.remove(wb['Sheet']) 
    else:
        wb = load_workbook(SUMMARY_REPORT_PATH)

    update_individual_summary_tab(wb, "Chase", today_str, c_adds, c_clears, c_91, c_0, c_3)
    update_individual_summary_tab(wb, "DB", today_str, d_adds, d_clears, d_91, d_0, d_3)
    wb.save(SUMMARY_REPORT_PATH)

    # 3. Update Accumulative Clears and Today's Clears Sheet
    update_accumulative_clears(c_clears_df, "Chase", cleared_date)
    update_accumulative_clears(d_clears_df, "DB", cleared_date)
    update_todays_clears_tab(c_clears_df, d_clears_df, cleared_date)

    # 4. Generate Import File (Adds)
    print("Generating FC_Import File...")
    combined_adds = pd.concat([
        format_adds_for_import(c_adds_df, 'Chase', today_str), 
        format_adds_for_import(d_adds_df, 'DB', today_str)
    ], ignore_index=True)

    if not combined_adds.empty:
        import_filename = f"FC_Import_{today.strftime('%m%d%Y')}.xlsx"
        import_path = os.path.join(IMPORT_FOLDER_PATH, import_filename)
        combined_adds.to_excel(import_path, index=False)
        print(f"Success: Created {import_filename} with {len(combined_adds)} rows.")
    else:
        print("No new adds found for today.")

    print("--- Process Complete ---")

if __name__ == "__main__":
    main()
