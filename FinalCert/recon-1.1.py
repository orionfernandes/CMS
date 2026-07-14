import pandas as pd
import os
import glob
import shutil
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.exceptions import InvalidFileException

# --- Configuration Variables ---
CHASE_FOLDER = r'c:/Users/OFernandes/Downloads/New folder (2)/New folder (2)/Daily Pipelines'
DB_FOLDER = r'c:/Users/OFernandes/Downloads/New folder (2)/New folder (2)/DB Pipelines'
SUMMARY_REPORT_PATH = r'C:/Users/OFernandes/Downloads/New folder (2)/New folder (2)/test/Summary_Report.xlsx'
IMPORT_FOLDER_PATH = r'C:/Users/OFernandes/Downloads/New folder (2)/New folder (2)/test/FC Imports' 
ACCUMULATIVE_BASE_DIR = r'C:/Users/OFernandes/Downloads/New folder (2)/New folder (2)/test/Accumulative_Clears'
BACKUP_DIR = os.path.join(ACCUMULATIVE_BASE_DIR, 'Backups')

os.makedirs(ACCUMULATIVE_BASE_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)
os.makedirs(IMPORT_FOLDER_PATH, exist_ok=True)

def format_all_dates(df):
    """Ensures dates are formatted as mm/dd/yyyy strings for Excel consistency."""
    for col in df.columns:
        if any(kw in str(col).lower() for kw in ['date', 'settlement']):
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%m/%d/%Y')
    return df

def get_target_files(folder_path, prefix, today_date):
    """Gathers all matching files, sorts them chronologically, and returns the 

    latest two files along with the latest file's parsed calendar date.
    """
    search_pattern = os.path.join(folder_path, f"{prefix}*.xlsx")
    all_files = glob.glob(search_pattern)
    valid_files = []
    for file in all_files:
        basename = os.path.basename(file)
        date_str = basename.replace(prefix, "").replace(".xlsx", "").strip()
        
        if date_str.startswith('_'):
            date_str = date_str[1:]
            
        f_date = None
        for fmt in ['%b.%d', '%B.%d']:
            try:
                f_date = datetime.strptime(date_str, fmt).replace(year=today_date.year)
                break
            except ValueError:
                continue
                
        if f_date is None: 
            continue
            
        if f_date > today_date + timedelta(days=30): 
            f_date = f_date.replace(year=today_date.year - 1)
        valid_files.append((f_date, file))
    
    # Sort files chronologically by their parsed calendar dates
    valid_files.sort(key=lambda x: x[0])
    if not valid_files: 
        return None, None, None
    
    # Pull absolute latest and second latest available files
    latest_date, latest_file = valid_files[-1]
    prev_file = valid_files[-2][1] if len(valid_files) > 1 else None
    
    return latest_file, prev_file, latest_date

def get_adds_and_clears(today_file, previous_file, sheet_name=0):
    try:
        df_today = pd.read_excel(today_file, sheet_name=sheet_name)
        df_prev = pd.read_excel(previous_file, sheet_name=sheet_name)
        df_today.columns = df_today.columns.str.strip()
        df_prev.columns = df_prev.columns.str.strip()
        df_today['ID_KEY'] = df_today['Unique ID'].astype(str).str.strip()
        df_prev['ID_KEY'] = df_prev['Unique ID'].astype(str).str.strip()
        
        today_ids = set(df_today['ID_KEY'].dropna())
        prev_ids = set(df_prev['ID_KEY'].dropna())
        
        adds_df = df_today[df_today['ID_KEY'].isin(today_ids - prev_ids)].copy()
        clears_df = df_prev[df_prev['ID_KEY'].isin(prev_ids - today_ids)].copy()
        
        # Calculate DQ category counts from today's file
        dq_91_180, dq_0_90, dq_3_yr = 0, 0, 0
        if 'Pool Age Status' in df_today.columns:
            pool_status = df_today['Pool Age Status'].astype(str).str.strip()
            dq_91_180 = len(pool_status[pool_status == 'DQ 91-180'])
            dq_0_90 = len(pool_status[pool_status == 'DQ 0-90'])
            dq_3_yr = len(pool_status[pool_status == '3yr DQ'])
        adds_df.drop(columns=['ID_KEY'], inplace=True, errors='ignore')
        clears_df.drop(columns=['ID_KEY'], inplace=True, errors='ignore')
        
        return len(adds_df), len(clears_df), adds_df, clears_df, dq_91_180, dq_0_90, dq_3_yr
    except Exception as e:
        print(f"Error comparing files: {e}")
        return 0, 0, pd.DataFrame(), pd.DataFrame(), 0, 0, 0

def update_accumulative_clears(clears_df, sheet_name, cleared_date_obj):
    if clears_df.empty: return
    cleared_date_str = cleared_date_obj.strftime('%m/%d/%Y')
    month_name = cleared_date_obj.strftime('%B_%Y') 
    file_path = os.path.join(ACCUMULATIVE_BASE_DIR, f"FC_Clear_Accumulative_{month_name}.xlsx")
    if os.path.exists(file_path):
        shutil.copy2(file_path, os.path.join(BACKUP_DIR, f"Backup_{os.path.basename(file_path)}"))
    clears_df = clears_df.copy()
    clears_df.insert(0, 'Cleared Date', cleared_date_str)
    clears_df = format_all_dates(clears_df)
    if not os.path.exists(file_path):
        wb = Workbook(); wb.active.title = "Chase"; wb.create_sheet("DB"); wb.save(file_path)
    
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames: wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    existing_records = set()
    if ws.max_row > 1:
        headers = [str(cell.value) for cell in ws[1]]
        try:
            date_idx = headers.index('Cleared Date')
            id_idx = headers.index('Unique ID')
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[date_idx] and row[id_idx]:
                    existing_records.add((str(row[date_idx]), str(row[id_idx])))
        except ValueError: pass
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        for col_num, header in enumerate(clears_df.columns, 1):
            ws.cell(1, col_num, header).font = Font(bold=True)
    added_count = 0
    for row_dict in clears_df.to_dict('records'):
        record_key = (str(row_dict['Cleared Date']), str(row_dict['Unique ID']))
        if record_key not in existing_records:
            ws.append(list(row_dict.values()))
            added_count += 1
    
    wb.save(file_path)
    print(f"Logged {added_count} new clears to {sheet_name} ({month_name} file).")

def update_todays_clears_tab(c_clears_df, d_clears_df, cleared_date_obj):
    """Creates a snapshot tab of all combined clears for the current run."""
    month_name = cleared_date_obj.strftime('%B_%Y')
    file_path = os.path.join(ACCUMULATIVE_BASE_DIR, f"FC_Clear_Accumulative_{month_name}.xlsx")
    if not os.path.exists(file_path):
        return
    c_df = c_clears_df.copy()
    d_df = d_clears_df.copy()
    if not c_df.empty:
        c_df.insert(0, 'Custodian', 'Chase')
        c_df.insert(0, 'Cleared Date', cleared_date_obj.strftime('%m/%d/%Y'))
        c_df = format_all_dates(c_df)
        
    if not d_df.empty:
        d_df.insert(0, 'Custodian', 'DB')
        d_df.insert(0, 'Cleared Date', cleared_date_obj.strftime('%m/%d/%Y'))
        d_df = format_all_dates(d_df)
    combined_clears = pd.concat([c_df, d_df], ignore_index=True)
    wb = load_workbook(file_path)
    sheet_name = "Today's Clears"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    if not combined_clears.empty:
        for col_num, header in enumerate(combined_clears.columns, 1):
            ws.cell(1, col_num, header).font = Font(bold=True)
        for row in combined_clears.values.tolist():
            ws.append(row)
            
    wb.save(file_path)
    # Today's Clears tab updated silently; main() prints the simple clear counts.

def format_adds_for_import(adds_df, client_type, exception_date_str):
    headers = [
        'Unique ID', 'Loan Number', 'Collateral Key', 'Settlement Date', 'Pool ID', 'Loan Status', 
        'Repool Investor', 'Prior Loan Number', 'Balance Principal', 'Responsible Party', 
        'Document Exception Status', 'Doctype Code', 'Doctype Descr', 'Doc Notation', 
        'Doc Status', 'Question Code', 'Question Descr', 'Question Notation', 
        'Exception Date', 'Clients Deal Name', 'Inv_#', 'Inv_Name'
    ]
    output_df = pd.DataFrame(columns=headers)
    if adds_df.empty: return output_df
    adds_df = format_all_dates(adds_df)
    if client_type == 'Chase':
        output_df['Unique ID'] = adds_df.get('Unique ID')
        output_df['Loan Number'] = adds_df.get('Collateral Key')
        output_df['Collateral Key'] = adds_df.get('Collateral Key')
        output_df['Settlement Date'] = adds_df.get('Settlement Date')
        output_df['Pool ID'] = adds_df.get('Pool Number')
        output_df['Loan Status'] = adds_df.get('Loan Status Current')
        output_df['Repool Investor'] = adds_df.get('Investor Name')
        output_df['Prior Loan Number'] = adds_df.get('Alternate Id')
        output_df['Balance Principal'] = adds_df.get('UPB') 
        output_df['Doc Status'] = adds_df.get('Doc Status')
        output_df['Inv_#'] = adds_df.get('Inv Num')
        output_df['Inv_Name'] = adds_df.get('Investor Name')
        output_df['Clients Deal Name'] = 'Chase'
    elif client_type == 'DB':
        output_df['Unique ID'] = adds_df.get('Unique ID')
        output_df['Loan Number'] = adds_df.get('Investor Collateral Key')
        output_df['Collateral Key'] = adds_df.get('Collateral Key')
        output_df['Settlement Date'] = adds_df.get('Settlement Date')
        output_df['Pool ID'] = adds_df.get('DB Pool Key')
        output_df['Loan Status'] = adds_df.get('Loan Status Current')
        output_df['Repool Investor'] = adds_df.get('Repool Investor')
        output_df['Prior Loan Number'] = adds_df.get('Prior Loan Number')
        output_df['Balance Principal'] = adds_df.get('Balance Principal')
        
        raw_status = adds_df.get('Document Condition Code')
        if raw_status is not None and hasattr(raw_status, 'map'):
            doc_status_mapping = {
                'ABL': "Attorney's Bailee Letter", 'PCR': 'Photocopy Recorded', 'UNR': 'Unrecorded',
                'P': 'Photo-Copy', 'O*': 'Original with comment', 'O': 'Original', 'M': 'Not Received',
                'NA': 'Not Applicable', 'LWC': 'LNA With A Copy of The Note', 'INC': 'Incomplete/Incorrect',
                'LNA': 'Lost Note Affidavit', 'CU': 'Copy - Certified by Unknown', 'EXT': 'Extra',
                'CC': 'Copy - County Certified', 'CS': 'Copy - Certified by Servicer/Seller', 'BLK': 'to Blank',
                'CMT': 'Commitment', 'PTR': 'Preliminary Title Report', 'ICM': 'Image - Title Commitment',
                'CT': 'Copy - Certified by Title Co.', 'ICS': 'Image - Copy Certified by Seller/Servicr',
                'CA': 'Copy - Certified By Attorney', 'CE': 'Copy - Certified by Escrow Co.',
                'IPR': 'Image - Preliminary Title Report', 'DNE': 'Document Unexecuted',
                'ICU': 'Image - Copy Certified by Unknown', 'ICT': 'Image - Copy Certified by Title Company',
                'SBL': 'Servicing Bailee Letter', 'AO': "Attorney's Opinion"
            }
            output_df['Doc Status'] = raw_status.astype(str).str.strip().map(doc_status_mapping).fillna(raw_status)
        else:
            output_df['Doc Status'] = raw_status
        
        output_df['Inv_#'] = adds_df.get('Investor Number')
        output_df['Inv_Name'] = adds_df.get('Investor Name')
        output_df['Clients Deal Name'] = 'DB'
    for field in ['Doctype Code', 'Doctype Descr', 'Doc Notation', 'Question Code', 'Question Descr', 'Question Notation']:
        output_df[field] = adds_df.get(field)
    
    output_df['Responsible Party'] = 'Final Certification'
    output_df['Document Exception Status'] = 'Analyst Pending Review'
    output_df['Exception Date'] = exception_date_str
    
    return output_df

def setup_summary_sheet_headers(ws, sheet_name):
    """Configures the precise multi-row layout of the Summary sheet tabs."""
    ws.cell(1, 1, f'{sheet_name}')
    ws.merge_cells('A1:G1')
    ws.cell(2, 1, 'Date')
    ws.cell(2, 2, 'Adds')
    ws.cell(2, 3, 'Clears')
    ws.cell(2, 4, 'Exception Count')
    ws.cell(2, 5, 'DQ')
    ws.cell(3, 5, '91-180')
    ws.cell(3, 6, '0-90')
    ws.cell(3, 7, '3 Year')
    ws.merge_cells('A2:A3')
    ws.merge_cells('B2:B3')
    ws.merge_cells('C2:C3')
    ws.merge_cells('D2:D3')
    ws.merge_cells('E2:G2')
    for row in [1, 2, 3]:
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)

def update_individual_summary_tab(wb, sheet_name, today_str, adds, clears, dq_91, dq_0, dq_3):
    """Updates an individual sheet inside the Summary Workbook."""
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        setup_summary_sheet_headers(ws, sheet_name)
    else:
        ws = wb[sheet_name]
        if ws.max_row == 1 and ws.cell(1,1).value is None:
            setup_summary_sheet_headers(ws, sheet_name)
    for row_idx in range(4, ws.max_row + 1):
        if str(ws.cell(row_idx, 1).value) == today_str:
            ws.delete_rows(row_idx)
            break
            
    last_row = ws.max_row
    prev_exc = int(ws.cell(last_row, 4).value or 0) if last_row >= 4 else 0
    curr_exc = prev_exc + adds - clears
    ws.append([today_str, adds, clears, curr_exc, dq_91, dq_0, dq_3])

def main():
    today = datetime.now()
    today_str = today.strftime('%m/%d/%Y')
    print(f"--- Process Start: {today_str} ---")
    
    # 1. Compare Files and Count DQs (Chase)
    c_today, c_prev, c_file_date = get_target_files(CHASE_FOLDER, 'FCERT_Pipeline_Loan_Detail_', today)
    if c_today and c_prev:
        c_adds, c_clears, c_adds_df, c_clears_df, c_91, c_0, c_3 = get_adds_and_clears(c_today, c_prev, 'Loan Detail')
        c_import_date = c_file_date.strftime('%m/%d/%Y')
        # Lookback based on file timestamp (If latest file is Monday, clear is Friday)
        c_cleared_date_obj = c_file_date - timedelta(days=3 if c_file_date.weekday() == 0 else 1)
    else:
        c_adds, c_clears, c_adds_df, c_clears_df, c_91, c_0, c_3 = (0,0,pd.DataFrame(),pd.DataFrame(),0,0,0)
        c_import_date = today_str
        c_cleared_date_obj = today - timedelta(days=3 if today.weekday() == 0 else 1)
        
    # 1b. Compare Files and Count DQs (DB)
    d_today, d_prev, d_file_date = get_target_files(DB_FOLDER, 'DB_FCERT_Pipeline_', today)
    if d_today and d_prev:
        d_adds, d_clears, d_adds_df, d_clears_df, d_91, d_0, d_3 = get_adds_and_clears(d_today, d_prev)
        d_import_date = d_file_date.strftime('%m/%d/%Y')
        # Lookback based on file timestamp (If latest file is Monday, clear is Friday)
        d_cleared_date_obj = d_file_date - timedelta(days=3 if d_file_date.weekday() == 0 else 1)
    else:
        d_adds, d_clears, d_adds_df, d_clears_df, d_91, d_0, d_3 = (0,0,pd.DataFrame(),pd.DataFrame(),0,0,0)
        d_import_date = today_str
        d_cleared_date_obj = today - timedelta(days=3 if today.weekday() == 0 else 1)
        
    # 2. Display simple clear counts
    print(f"Chase clears found: {c_clears}")
    print(f"DB clears found: {d_clears}")
    print(f"Total clears logged today: {c_clears + d_clears}")

    # 3. Update New Tabbed Summary Report
    print("Updating Separate Summary Tabs...")
    if not os.path.exists(SUMMARY_REPORT_PATH):
        wb = Workbook()
        if 'Sheet' in wb.sheetnames: wb.remove(wb['Sheet']) 
    else:
        wb = load_workbook(SUMMARY_REPORT_PATH)
    update_individual_summary_tab(wb, "Chase", today_str, c_adds, c_clears, c_91, c_0, c_3)
    update_individual_summary_tab(wb, "DB", today_str, d_adds, d_clears, d_91, d_0, d_3)
    wb.save(SUMMARY_REPORT_PATH)
    
    # 4. Update Accumulative Clears and Today's Clears Sheet
    update_accumulative_clears(c_clears_df, "Chase", c_cleared_date_obj)
    update_accumulative_clears(d_clears_df, "DB", d_cleared_date_obj)
    update_todays_clears_tab(c_clears_df, d_clears_df, c_cleared_date_obj)
    
    # 5. Generate Import File (Adds) using parsed file dates
    print("Generating FC_Import File...")
    combined_adds = pd.concat([
        format_adds_for_import(c_adds_df, 'Chase', c_import_date), 
        format_adds_for_import(d_adds_df, 'DB', d_import_date)
    ], ignore_index=True)
    if not combined_adds.empty:
        import_filename = f"FC_Import_{today.strftime('%m%d%Y')}.xlsx"
        import_path = os.path.join(IMPORT_FOLDER_PATH, import_filename)
        combined_adds.to_excel(import_path, index=False)
        print(f"Success: Created {import_filename} with {len(combined_adds)} rows.")
    else:
        print("No new adds found for today.")
    print("--- Process Complete ---")

if __name__ == "__main__":
    main()
